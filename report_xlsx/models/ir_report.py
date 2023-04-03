# -*- coding: utf-8 -*-
import os
import io
from io import BytesIO
import base64
import re

from odoo import api, fields, models, _
from odoo.exceptions import UserError, Warning

import openpyxl
from openpyxl.cell.cell import *
import openpyxl.styles  as ExStyle
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

import logging
import traceback
_logger = logging.getLogger(__name__)

def insert_rows(self, row_idx, cnt=1, above=False, copy_style=True, copy_merged_columns=True, fill_formulae=True):
    """Inserts new (empty) rows into worksheet at specified row index.

    :param row_idx: Row index specifying where to insert new rows.
    :param cnt: Number of rows to insert.
    :param above: Set True to insert rows above specified row index.
    :param copy_style: Set True if new rows should copy style of immediately above row.
    :param fill_formulae: Set True if new rows should take on formula from immediately above row, filled with references new to rows.

    Usage:
    * insert_rows(2, 10, above=True, copy_style=False)
    """
    CELL_RE  = re.compile("(?P<col>\$?[A-Z]+)(?P<row>\$?\d+)")

    row_idx = row_idx - 1 if above else row_idx
    def replace(m):
        row = m.group('row')
        prefix = "$" if row.find("$") != -1 else ""
        row = int(row.replace("$", ""))
        row += cnt if row > row_idx else 0
        return m.group('col') + prefix + str(row)

    # First, we shift all cells down cnt rows...
    old_cells = set()
    old_fas = set()
    new_cells = dict()
    new_fas = dict()
    for c in self._cells.values():
        old_coor = c.coordinate

        # Shift all references to anything below row_idx
        if c.data_type == TYPE_FORMULA:
            c.value = CELL_RE.sub(
                replace,
                c.value
            )
            # Here, we need to properly update the formula references to reflect new row indices
            if old_coor in self.formula_attributes and 'ref' in self.formula_attributes[old_coor]:
                self.formula_attributes[old_coor]['ref'] = CELL_RE.sub(
                    replace,
                    self.formula_attributes[old_coor]['ref']
                )

        # Do the magic to set up our actual shift
        if c.row > row_idx:
            old_coor = c.coordinate
            old_cells.add((c.row, c.column))
            c.row += cnt
            new_cells[(c.row, c.column)] = c
            if old_coor in self.formula_attributes:
                old_fas.add(old_coor)
                fa = self.formula_attributes[old_coor].copy()
                new_fas[c.coordinate] = fa

    for coor in old_cells:
        del self._cells[coor]
    self._cells.update(new_cells)

    for fa in old_fas:
        del self.formula_attributes[fa]
    self.formula_attributes.update(new_fas)

    # Next, we need to shift all the Row Dimensions below our new rows down by cnt...
    import copy
    if list(self.row_dimensions):
        for row in range(list(self.row_dimensions)[-1] + cnt, row_idx + cnt, -1):
            new_rd = copy.copy(self.row_dimensions[row - cnt])
            new_rd.index = row
            self.row_dimensions[row] = new_rd
            del self.row_dimensions[row - cnt]

    # Now, create our new rows, with all the pretty cells
    new_row_idx = row_idx + 1
    for row in range(new_row_idx, new_row_idx + cnt):
        # Create a Row Dimension for our new row
        new_rd = copy.copy(self.row_dimensions[row-1])
        new_rd.index = row
        self.row_dimensions[row] = new_rd

        for col in range(self.max_column):
            col2 = get_column_letter(col+1)
            cell = self.cell(row, col+1)
            cell.value = None
            source = self.cell(row-1, col+1)
            source_copy = self.cell(row-2, col+1,)
            if copy_style:
                cell.number_format = source.number_format
                cell.font = source.font.copy()
                cell.alignment = source.alignment.copy()
                cell.border = source.border.copy()
                cell.fill = source.fill.copy()
            if fill_formulae and source.data_type == TYPE_FORMULA:
                s_coor = source.coordinate
                if s_coor in self.formula_attributes and 'ref' not in self.formula_attributes[s_coor]:
                    fa = self.formula_attributes[s_coor].copy()
                    self.formula_attributes[cell.coordinate] = fa
                #print("Copying formula from cell %s%d to %s%d"%(col,row-1,col,row))
                cell.value = re.sub(
                    "(\$?[A-Z]{1,3}\$?)%d" % (row-1),
                    lambda m: (m.group(1) + str(row)) if m.group(2) == str(row-1) else (m.group(1) + m.group(2)),
                    source.value
                )
                cell.data_type = TYPE_FORMULA
            source.number_format = source_copy.number_format
            source.font = source_copy.font.copy()
            source.alignment = source_copy.alignment.copy()
            source.border = source_copy.border.copy()
            source.fill = source_copy.fill.copy()

    # Check for Merged Cell Ranges that need to be expanded to contain new cells
    for cr_idx, cr in enumerate(self.merged_cells.ranges[:]):
        self.merged_cells.ranges[cr_idx] = CELL_RE.sub(
            replace,
            str(cr)
        )

    # Merge columns of the new rows in the same way row above does
    if copy_merged_columns:
        for cr in self.merged_cell_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(cr)
            if max_row == min_row == row_idx:
                for row in range(new_row_idx, new_row_idx + cnt):
                    newCellRange = get_column_letter(min_col) + str(row) + ":" + get_column_letter(max_col) + str(row)
                    self.merge_cells(newCellRange)

Worksheet.insert_rows = insert_rows


class ReportAction(models.Model):
    _inherit = "ir.actions.report"

    report_type = fields.Selection(selection_add=[("xlsx", "XLSX")], ondelete={'xlsx': lambda recs: recs.write({'report_type': 'txt'})})
    excel_attachment_id = fields.Many2one("ir.attachment", "Excel模板")
    code = fields.Text(string='Python Code', default="""# 
#Variant you can use
#'env': self.env,
#'Warning': odoo.exceptions.Warning,
#'obj': objs,
#'ws': ws,
#'style': ExStyle,
#'add_img': add_img,
#'log': _logger,
#'traceback': traceback,
#
#Example:
#ws['G9'] = obj.confirmation_date
#ws["B1"].font=style.Font(name="宋体"，size=17,color="00CCFF")
#ws.cell(2,1).fill = style.PatternFill("solid",fgColor="00FF02")
#ws.cell(row,col).alignment = style.Alignment(horizontal='center', vertical='justify', wrap_text=True)
#ws.merge_cells('A1:D4')
#ws.unmerge_cells('A1:D4')
#row = 17
#idx = 1
#for line in obj.order_line:
#    ws.cell(row,1).value = idx
#    add_img(ws, (row,2), line.product_id.image_medium, (100, 100))
#    ws.cell(row,3).value = line.product_id.with_context(lang=obj.partner_id.lang).name
#    ws.insert_rows(row,1)
#
#    row=row+1
#    idx=idx+1
1==1""")

    @api.model
    def render_xlsx_eval_context(self, res_ids, data=None):
        def add_img(ws, anchor, img, size):
            from ._utils import image_resize_image
            img1 = img
            if size:
                img1 = image_resize_image(img, size)
            image_file = io.BytesIO(base64.b64decode(img1))
            img2 = openpyxl.drawing.image.Image(image_file)
            img2.anchor = "%s%s" % (get_column_letter(anchor[1]),anchor[0]) # A1
            ws.add_image(img2)
            
        eval_context = {}
        eval_context.update({
            # orm
            'env': self.env,
            # Exceptions
            'Warning': Warning,
            'style': ExStyle,
            #helper
            'add_img': add_img,
            'log': _logger,
            'traceback': traceback,
        })
        return eval_context

    @api.model
    def render_xlsx(self, res_ids, data=None):
        if not self.excel_attachment_id:
            raise UserError("No Excel Report Template Attachment, Please set first!")

        if not data:
            data = {}
        data.setdefault('report_type', 'xlsx')
        self_sudo = self.sudo()
        record_ids = []
        if res_ids:
            Model = self.env[self_sudo.model]
            record_ids = Model.browse(res_ids)

        fp = BytesIO(base64.b64decode(self.excel_attachment_id.datas))
        wb = openpyxl.load_workbook(fp)
        ws = wb.active
        eval_context = self.render_xlsx_eval_context(res_ids, data)
        eval_context.update({
            # record
            'obj': record_ids,
            'wb': wb,
            'ws': ws,
            })

        for obj in record_ids:
            if len(record_ids) > 1:
                ws = wb.copy_worksheet(ws)
            else:
                ws = ws
            ws.title = ws.title or "Sheet"

            eval_context['obj'] = obj
            eval_context['ws'] = ws
            exec(self.code.strip(), eval_context)

        fdata = save_virtual_workbook(wb)
        return fdata, 'xlsx'
